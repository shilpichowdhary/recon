"""Excel report generator for reconciliation results."""

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.performance import PortfolioReconciliation, ReconciliationResult, ReconciliationStatus
from models.transaction import Transaction
from calculators.pnl_calculator import PortfolioPnL
from .formatters import ExcelFormatter


class ExcelReportGenerator:
    """
    Generate comprehensive Excel reconciliation reports.

    Creates a 7-sheet workbook:
    1. Executive Summary
    2. Performance Reconciliation
    3. P&L Reconciliation
    4. Position Detail (FIFO lots)
    5. Cash Flows
    6. Data Quality
    7. Audit Trail
    """

    def __init__(self, output_path: Optional[Path] = None):
        """
        Initialize report generator.

        Args:
            output_path: Path for output file
        """
        self.output_path = output_path or Path("reconciliation_report.xlsx")
        self.workbook: Optional[Workbook] = None
        self.formatter = ExcelFormatter()

    def generate_report(
        self,
        reconciliation: PortfolioReconciliation,
        transactions: List[Transaction],
        portfolio_pnl: PortfolioPnL,
        lot_details: Dict[str, List[Dict]],
        cash_flow_summary: Dict[str, Any],
        data_quality_issues: List[str],
    ) -> Path:
        """
        Generate complete reconciliation report.

        Args:
            reconciliation: Reconciliation results
            transactions: Transaction list
            portfolio_pnl: P&L calculation results
            lot_details: FIFO lot details by symbol
            cash_flow_summary: Cash flow summary
            data_quality_issues: List of data quality issues

        Returns:
            Path to generated report
        """
        self.workbook = Workbook()

        # Remove default sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # Create sheets
        self._create_executive_summary(reconciliation)
        self._create_performance_reconciliation(reconciliation)
        self._create_pnl_reconciliation(reconciliation, portfolio_pnl)
        self._create_position_detail(lot_details)
        self._create_cash_flows(transactions, cash_flow_summary)
        self._create_data_quality(data_quality_issues)
        self._create_audit_trail(reconciliation, transactions)

        # Save workbook
        self.workbook.save(self.output_path)

        return self.output_path

    def _create_executive_summary(self, recon: PortfolioReconciliation) -> None:
        """Create Executive Summary sheet."""
        ws = self.workbook.create_sheet("Executive Summary")

        # Title
        self.formatter.add_title(ws, "PMS Reconciliation Report", row=1)

        # Report metadata
        ws.cell(row=3, column=1, value="Portfolio ID:")
        ws.cell(row=3, column=2, value=recon.portfolio_id)

        ws.cell(row=4, column=1, value="Reconciliation Date:")
        ws.cell(row=4, column=2, value=recon.reconciliation_date.isoformat())

        ws.cell(row=5, column=1, value="Base Currency:")
        ws.cell(row=5, column=2, value=recon.base_currency)

        ws.cell(row=6, column=1, value="Report Generated:")
        ws.cell(row=6, column=2, value=datetime.now().isoformat())

        # Summary statistics
        self.formatter.add_subtitle(ws, "Reconciliation Summary", row=8)

        summary = recon.get_summary()

        row = 10
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.formatter.format_header_row(ws, row, len(headers))

        summary_data = [
            ("Total Checks", summary["total_checks"]),
            ("Passed", summary["passed"]),
            ("Failed", summary["failed"]),
            ("Warnings", summary["warnings"]),
            ("Pass Rate", summary["pass_rate"]),
            ("Fully Reconciled", "Yes" if summary["fully_reconciled"] else "No"),
            ("Data Quality Issues", summary["data_quality_issues"]),
        ]

        for i, (metric, value) in enumerate(summary_data, 1):
            ws.cell(row=row + i, column=1, value=metric)
            ws.cell(row=row + i, column=2, value=value)
            self.formatter.format_data_row(ws, row + i, 2, alternate=(i % 2 == 0))

        # Status indicator
        status_row = row + len(summary_data) + 2
        ws.cell(row=status_row, column=1, value="Overall Status:")

        status_cell = ws.cell(row=status_row, column=2)
        if recon.is_fully_reconciled:
            status_cell.value = "PASS"
            self.formatter.format_status_cell(ws, status_row, 2, "PASS")
        else:
            status_cell.value = "FAIL"
            self.formatter.format_status_cell(ws, status_row, 2, "FAIL")

        # Failed items (if any)
        if recon.failed_results:
            self.formatter.add_subtitle(ws, "Failed Reconciliation Items", row=status_row + 2)

            fail_headers = ["Metric", "Symbol", "Calculated", "Expected", "Difference"]
            fail_row = status_row + 4
            for col, header in enumerate(fail_headers, 1):
                ws.cell(row=fail_row, column=col, value=header)
            self.formatter.format_header_row(ws, fail_row, len(fail_headers))

            for i, result in enumerate(recon.failed_results[:10], 1):  # Limit to 10
                ws.cell(row=fail_row + i, column=1, value=result.metric_name)
                ws.cell(row=fail_row + i, column=2, value=result.symbol or "Portfolio")
                ws.cell(row=fail_row + i, column=3, value=float(result.calculated_value))
                ws.cell(row=fail_row + i, column=4, value=float(result.expected_value))
                ws.cell(row=fail_row + i, column=5, value=float(result.difference))
                self.formatter.format_data_row(ws, fail_row + i, len(fail_headers))

        self.formatter.auto_fit_columns(ws)

    def _create_performance_reconciliation(self, recon: PortfolioReconciliation) -> None:
        """Create Performance Reconciliation sheet."""
        ws = self.workbook.create_sheet("Performance Reconciliation")

        self.formatter.add_title(ws, "Performance Metrics Reconciliation", row=1)

        # Filter to performance metrics
        perf_metrics = ["XIRR", "IRR", "TWR", "TWR Annualized"]
        perf_results = [r for r in recon.results if r.metric_name in perf_metrics]

        headers = ["Metric", "Calculated", "Expected", "Difference", "Tolerance", "Status"]
        row = 3

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.formatter.format_header_row(ws, row, len(headers))

        for i, result in enumerate(perf_results, 1):
            data_row = row + i
            ws.cell(row=data_row, column=1, value=result.metric_name)
            ws.cell(row=data_row, column=2, value=f"{float(result.calculated_value):.4%}")
            ws.cell(row=data_row, column=3, value=f"{float(result.expected_value):.4%}")
            ws.cell(row=data_row, column=4, value=f"{float(result.difference):.4%}")
            ws.cell(row=data_row, column=5, value=f"{float(result.tolerance):.4%}")
            ws.cell(row=data_row, column=6, value=result.status.value)

            self.formatter.format_data_row(ws, data_row, len(headers), alternate=(i % 2 == 0))
            self.formatter.format_status_cell(ws, data_row, 6, result.status.value)

        # Add calculated metrics if available
        if recon.calculated_metrics:
            metrics = recon.calculated_metrics
            self.formatter.add_subtitle(ws, "Calculated Performance Metrics", row=row + len(perf_results) + 3)

            calc_row = row + len(perf_results) + 5
            calc_data = [
                ("XIRR", f"{float(metrics.xirr):.4%}" if metrics.xirr else "N/A"),
                ("TWR", f"{float(metrics.twr):.4%}" if metrics.twr else "N/A"),
                ("TWR Annualized", f"{float(metrics.twr_annualized):.4%}" if metrics.twr_annualized else "N/A"),
                ("Period Start", metrics.start_date.isoformat() if metrics.start_date else "N/A"),
                ("Period End", metrics.end_date.isoformat() if metrics.end_date else "N/A"),
            ]

            for i, (label, value) in enumerate(calc_data):
                ws.cell(row=calc_row + i, column=1, value=label)
                ws.cell(row=calc_row + i, column=2, value=value)

        self.formatter.auto_fit_columns(ws)

    def _create_pnl_reconciliation(self, recon: PortfolioReconciliation, pnl: PortfolioPnL) -> None:
        """Create P&L Reconciliation sheet."""
        ws = self.workbook.create_sheet("P&L Reconciliation")

        self.formatter.add_title(ws, "P&L Reconciliation", row=1)

        # Portfolio level P&L
        self.formatter.add_subtitle(ws, "Portfolio Summary", row=3)

        headers = ["Metric", "Calculated", "Expected", "Difference", "Status"]
        row = 5

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.formatter.format_header_row(ws, row, len(headers))

        # Filter P&L related results
        pnl_metrics = ["Realized P&L", "Unrealized P&L", "Total P&L", "Market Value", "Cost Basis"]
        pnl_results = [r for r in recon.results if r.metric_name in pnl_metrics and r.symbol is None]

        for i, result in enumerate(pnl_results, 1):
            data_row = row + i
            ws.cell(row=data_row, column=1, value=result.metric_name)
            ws.cell(row=data_row, column=2, value=f"${float(result.calculated_value):,.2f}")
            ws.cell(row=data_row, column=3, value=f"${float(result.expected_value):,.2f}")
            ws.cell(row=data_row, column=4, value=f"${float(result.difference):,.2f}")
            ws.cell(row=data_row, column=5, value=result.status.value)

            self.formatter.format_data_row(ws, data_row, len(headers), alternate=(i % 2 == 0))
            self.formatter.format_status_cell(ws, data_row, 5, result.status.value)

        # Position level P&L
        pos_row = row + len(pnl_results) + 3
        self.formatter.add_subtitle(ws, "Position P&L Details", row=pos_row)

        pos_headers = ["Symbol", "Quantity", "Cost Basis", "Market Value", "Realized P&L", "Unrealized P&L", "Total P&L"]
        pos_row += 2

        for col, header in enumerate(pos_headers, 1):
            ws.cell(row=pos_row, column=col, value=header)
        self.formatter.format_header_row(ws, pos_row, len(pos_headers))

        for i, (symbol, position) in enumerate(pnl.positions.items(), 1):
            data_row = pos_row + i
            ws.cell(row=data_row, column=1, value=symbol)
            ws.cell(row=data_row, column=2, value=float(position.quantity))
            ws.cell(row=data_row, column=3, value=float(position.cost_basis))
            ws.cell(row=data_row, column=4, value=float(position.market_value))
            ws.cell(row=data_row, column=5, value=float(position.realized_pnl))
            ws.cell(row=data_row, column=6, value=float(position.unrealized_pnl))
            ws.cell(row=data_row, column=7, value=float(position.total_pnl))

            self.formatter.format_data_row(ws, data_row, len(pos_headers), alternate=(i % 2 == 0))

        self.formatter.auto_fit_columns(ws)

    def _create_position_detail(self, lot_details: Dict[str, List[Dict]]) -> None:
        """Create Position Detail (FIFO lots) sheet."""
        ws = self.workbook.create_sheet("Position Detail")

        self.formatter.add_title(ws, "Position Detail - FIFO Lot Tracking", row=1)

        headers = [
            "Symbol", "Lot ID", "Acquisition Date", "Acquisition Price",
            "Original Qty", "Remaining Qty", "Cost Basis", "Cost/Unit",
            "Holding Days", "Term"
        ]
        row = 3

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.formatter.format_header_row(ws, row, len(headers))

        data_row = row + 1
        for symbol, lots in lot_details.items():
            for i, lot in enumerate(lots):
                ws.cell(row=data_row, column=1, value=symbol)
                ws.cell(row=data_row, column=2, value=lot.get("lot_id", "")[:8])
                ws.cell(row=data_row, column=3, value=lot.get("acquisition_date", ""))
                ws.cell(row=data_row, column=4, value=lot.get("acquisition_price", 0))
                ws.cell(row=data_row, column=5, value=lot.get("acquisition_quantity", lot.get("original_quantity", 0)))
                ws.cell(row=data_row, column=6, value=lot.get("remaining_quantity", 0))
                ws.cell(row=data_row, column=7, value=lot.get("cost_basis", 0))
                ws.cell(row=data_row, column=8, value=lot.get("cost_per_unit", 0))
                ws.cell(row=data_row, column=9, value=lot.get("holding_days", 0))
                ws.cell(row=data_row, column=10, value=lot.get("holding_period", ""))

                self.formatter.format_data_row(ws, data_row, len(headers), alternate=(data_row % 2 == 0))
                data_row += 1

        self.formatter.auto_fit_columns(ws)

    def _create_cash_flows(self, transactions: List[Transaction], summary: Dict[str, Any]) -> None:
        """Create Cash Flows sheet."""
        ws = self.workbook.create_sheet("Cash Flows")

        self.formatter.add_title(ws, "Cash Flow Analysis", row=1)

        # Summary
        self.formatter.add_subtitle(ws, "Summary", row=3)

        summary_data = [
            ("Total Inflows", f"${summary.get('total_inflows', 0):,.2f}"),
            ("Total Outflows", f"${summary.get('total_outflows', 0):,.2f}"),
            ("Total Income", f"${summary.get('total_income', 0):,.2f}"),
            ("Net Cash Flow", f"${summary.get('net_cash_flow', 0):,.2f}"),
            ("Transaction Count", summary.get("transaction_count", 0)),
        ]

        row = 5
        for i, (label, value) in enumerate(summary_data):
            ws.cell(row=row + i, column=1, value=label)
            ws.cell(row=row + i, column=2, value=value)

        # Transaction detail
        detail_row = row + len(summary_data) + 2
        self.formatter.add_subtitle(ws, "Transaction Detail", row=detail_row)

        headers = ["Date", "Type", "Symbol", "Quantity", "Price", "Gross Amount", "Fees", "Net Amount", "Currency"]
        detail_row += 2

        for col, header in enumerate(headers, 1):
            ws.cell(row=detail_row, column=col, value=header)
        self.formatter.format_header_row(ws, detail_row, len(headers))

        # Sort transactions by date
        sorted_txns = sorted(transactions, key=lambda t: t.transaction_date)

        for i, txn in enumerate(sorted_txns, 1):
            data_row = detail_row + i
            ws.cell(row=data_row, column=1, value=txn.transaction_date.isoformat())
            ws.cell(row=data_row, column=2, value=txn.transaction_type.name)
            ws.cell(row=data_row, column=3, value=txn.symbol)
            ws.cell(row=data_row, column=4, value=float(txn.quantity))
            ws.cell(row=data_row, column=5, value=float(txn.price))
            ws.cell(row=data_row, column=6, value=float(txn.gross_amount))
            ws.cell(row=data_row, column=7, value=float(txn.total_fees))
            ws.cell(row=data_row, column=8, value=float(txn.net_amount))
            ws.cell(row=data_row, column=9, value=txn.currency.value)

            self.formatter.format_data_row(ws, data_row, len(headers), alternate=(i % 2 == 0))

        self.formatter.auto_fit_columns(ws)

    def _create_data_quality(self, issues: List[str]) -> None:
        """Create Data Quality sheet."""
        ws = self.workbook.create_sheet("Data Quality")

        self.formatter.add_title(ws, "Data Quality Report", row=1)

        ws.cell(row=3, column=1, value="Total Issues:")
        ws.cell(row=3, column=2, value=len(issues))

        if not issues:
            ws.cell(row=5, column=1, value="No data quality issues found.")
            self.formatter.format_status_cell(ws, 5, 1, "PASS")
        else:
            headers = ["#", "Issue Description"]
            row = 5

            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self.formatter.format_header_row(ws, row, len(headers))

            for i, issue in enumerate(issues, 1):
                data_row = row + i
                ws.cell(row=data_row, column=1, value=i)
                ws.cell(row=data_row, column=2, value=issue)
                self.formatter.format_data_row(ws, data_row, len(headers), alternate=(i % 2 == 0))

        self.formatter.auto_fit_columns(ws)

    def _create_audit_trail(self, recon: PortfolioReconciliation, transactions: List[Transaction]) -> None:
        """Create Audit Trail sheet."""
        ws = self.workbook.create_sheet("Audit Trail")

        self.formatter.add_title(ws, "Reconciliation Audit Trail", row=1)

        # Reconciliation metadata
        audit_data = [
            ("Report Generated", datetime.now().isoformat()),
            ("Portfolio ID", recon.portfolio_id),
            ("Reconciliation Date", recon.reconciliation_date.isoformat()),
            ("Base Currency", recon.base_currency),
            ("Total Transactions", len(transactions)),
            ("Total Checks Performed", recon.total_checks),
            ("Checks Passed", recon.passed_checks),
            ("Checks Failed", recon.failed_checks),
        ]

        row = 3
        for i, (label, value) in enumerate(audit_data):
            ws.cell(row=row + i, column=1, value=label)
            ws.cell(row=row + i, column=2, value=value)

        # All reconciliation results
        results_row = row + len(audit_data) + 2
        self.formatter.add_subtitle(ws, "All Reconciliation Results", row=results_row)

        headers = ["Metric", "Symbol", "Calculated", "Expected", "Difference", "Tolerance", "Status"]
        results_row += 2

        for col, header in enumerate(headers, 1):
            ws.cell(row=results_row, column=col, value=header)
        self.formatter.format_header_row(ws, results_row, len(headers))

        for i, result in enumerate(recon.results, 1):
            data_row = results_row + i
            ws.cell(row=data_row, column=1, value=result.metric_name)
            ws.cell(row=data_row, column=2, value=result.symbol or "Portfolio")
            ws.cell(row=data_row, column=3, value=float(result.calculated_value))
            ws.cell(row=data_row, column=4, value=float(result.expected_value))
            ws.cell(row=data_row, column=5, value=float(result.difference))
            ws.cell(row=data_row, column=6, value=float(result.tolerance))
            ws.cell(row=data_row, column=7, value=result.status.value)

            self.formatter.format_data_row(ws, data_row, len(headers), alternate=(i % 2 == 0))
            self.formatter.format_status_cell(ws, data_row, 7, result.status.value)

        self.formatter.auto_fit_columns(ws)
