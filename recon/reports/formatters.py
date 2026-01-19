"""Excel formatting utilities."""

from typing import Optional
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelFormatter:
    """Utility class for Excel formatting."""

    # Colors
    HEADER_BG = "4472C4"  # Blue
    PASS_BG = "C6EFCE"     # Green
    FAIL_BG = "FFC7CE"     # Red
    WARNING_BG = "FFEB9C"  # Yellow
    ALTERNATE_ROW = "F2F2F2"  # Light gray

    # Fonts
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(bold=True, size=12)
    NORMAL_FONT = Font(size=10)

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Alignments
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')

    @classmethod
    def get_header_fill(cls) -> PatternFill:
        """Get header background fill."""
        return PatternFill(start_color=cls.HEADER_BG, end_color=cls.HEADER_BG, fill_type="solid")

    @classmethod
    def get_pass_fill(cls) -> PatternFill:
        """Get pass status background fill."""
        return PatternFill(start_color=cls.PASS_BG, end_color=cls.PASS_BG, fill_type="solid")

    @classmethod
    def get_fail_fill(cls) -> PatternFill:
        """Get fail status background fill."""
        return PatternFill(start_color=cls.FAIL_BG, end_color=cls.FAIL_BG, fill_type="solid")

    @classmethod
    def get_warning_fill(cls) -> PatternFill:
        """Get warning status background fill."""
        return PatternFill(start_color=cls.WARNING_BG, end_color=cls.WARNING_BG, fill_type="solid")

    @classmethod
    def get_alternate_fill(cls) -> PatternFill:
        """Get alternate row background fill."""
        return PatternFill(start_color=cls.ALTERNATE_ROW, end_color=cls.ALTERNATE_ROW, fill_type="solid")

    @classmethod
    def format_header_row(cls, ws: Worksheet, row: int, num_cols: int) -> None:
        """Format a row as header."""
        header_fill = cls.get_header_fill()

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cls.HEADER_FONT
            cell.fill = header_fill
            cell.alignment = cls.CENTER
            cell.border = cls.THIN_BORDER

    @classmethod
    def format_data_row(cls, ws: Worksheet, row: int, num_cols: int, alternate: bool = False) -> None:
        """Format a data row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cls.NORMAL_FONT
            cell.border = cls.THIN_BORDER
            if alternate:
                cell.fill = cls.get_alternate_fill()

    @classmethod
    def format_status_cell(cls, ws: Worksheet, row: int, col: int, status: str) -> None:
        """Format a status cell with appropriate color."""
        cell = ws.cell(row=row, column=col)

        if status.upper() == "PASS":
            cell.fill = cls.get_pass_fill()
        elif status.upper() == "FAIL":
            cell.fill = cls.get_fail_fill()
        elif status.upper() == "WARNING":
            cell.fill = cls.get_warning_fill()

    @classmethod
    def set_column_widths(cls, ws: Worksheet, widths: dict) -> None:
        """
        Set column widths.

        Args:
            ws: Worksheet
            widths: Dict of column letter or index -> width
        """
        for col, width in widths.items():
            if isinstance(col, int):
                col = get_column_letter(col)
            ws.column_dimensions[col].width = width

    @classmethod
    def auto_fit_columns(cls, ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-fit column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def add_title(cls, ws: Worksheet, title: str, row: int = 1, col: int = 1) -> None:
        """Add a title to the worksheet."""
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = cls.TITLE_FONT

    @classmethod
    def add_subtitle(cls, ws: Worksheet, subtitle: str, row: int, col: int = 1) -> None:
        """Add a subtitle to the worksheet."""
        cell = ws.cell(row=row, column=col, value=subtitle)
        cell.font = cls.SUBTITLE_FONT

    @classmethod
    def format_currency(cls, value: float, decimals: int = 2) -> str:
        """Format a number as currency."""
        return f"${value:,.{decimals}f}"

    @classmethod
    def format_percentage(cls, value: float, decimals: int = 2) -> str:
        """Format a number as percentage."""
        return f"{value * 100:.{decimals}f}%"

    @classmethod
    def format_number(cls, value: float, decimals: int = 2) -> str:
        """Format a number with thousands separator."""
        return f"{value:,.{decimals}f}"

    @classmethod
    def create_named_styles(cls) -> dict:
        """Create named styles for the workbook."""
        styles = {}

        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '"$"#,##0.00'
        styles["currency"] = currency_style

        # Percentage style
        pct_style = NamedStyle(name="percentage")
        pct_style.number_format = '0.00%'
        styles["percentage"] = pct_style

        # Date style
        date_style = NamedStyle(name="date")
        date_style.number_format = 'YYYY-MM-DD'
        styles["date"] = date_style

        return styles
